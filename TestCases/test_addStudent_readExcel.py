import requests
import jsonpath
import json
import openpyxl


def test_add_multiple_students():
    APi_URL = "http://thetestingworldapi.com/api/studentsDetails"
    file = open("C:\\Users\\Chandu\\Desktop\\StudentData.json", 'r')
    json_request = json.loads(file.read())
    wb = openpyxl.load_workbook("C:\\Users\\Chandu\\Desktop\\TestData.xlsx")
    sh = wb['Sheet1']
    rows = sh.max_row
    for i in range(2, rows+1):
        first_name = sh.cell(row=i, column=1)
        mid_name = sh.cell(row=i, column=2)
        last_name = sh.cell(row=i, column=3)
        dob = sh.cell(row=i, column=4)
        json_request['first_name'] = first_name.value
        json_request['middle_name'] = mid_name.value
        json_request['last_name'] = last_name.value
        json_request['date_of_birth'] = dob.value
        response = requests.post(APi_URL, json_request)
        print(response.status_code)
        print(response.text)
        assert response.status_code == 201
