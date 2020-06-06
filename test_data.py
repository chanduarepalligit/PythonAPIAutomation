import openpyxl


class DataRead:

    def __init__(self, filePath, sheetName):
        global sheet
        wb = openpyxl.load_workbook(filePath)
        sheet = wb[sheetName]

    def column_count(self):
        cols = sheet.max_column
        return cols

    def row_count(self):
        rows = sheet.max_row
        return rows

    def fetch_key_Names(self):
        cols = sheet.max_column
        li = []
        for i in range(1, cols+1):
            cell = sheet.cell(row=1, column=i)
            li.insert(i, cell.value)
        return li

    def update_jsonRequest(self, rowNumber, jsonRequest, keyList):
        cols = sheet.max_column
        for i in range(1, cols+1):
            cell = sheet.cell(row=rowNumber, column=i)
            jsonRequest[keyList[i-1]] = cell.value
        return jsonRequest

