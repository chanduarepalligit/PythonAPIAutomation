import openpyxl


class DataRead:

    def __init__(self, FileName, SheetName):
        global sh
        wb = openpyxl.load_workbook(FileName)
        sh = wb[SheetName]

    def row_count(self):
        rows = sh.max_row
        return rows

    def column_count(self):
        columns = sh.max_column
        return columns

    def fetch_KeyNames(self):
        columns = sh.max_column
        li = []
        for i in range(1, columns + 1):
            cell = sh.cell(row=1, column=i)
            li.insert(i - 1, cell.value)
        return li

    def update_request_with_data(self, rowNumber, jsonRequest, keyList):
        columns = sh.max_column
        for i in range(1, columns + 1):
            c = sh.cell(row=rowNumber, column=i)
            jsonRequest[keyList[i-1]] = c.value
        return jsonRequest
